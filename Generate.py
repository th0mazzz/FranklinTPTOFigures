'''
****************** IMPORTING MODULES ******************
'''
import os
import math
import pandas as pd
import openpyxl
import xlsxwriter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from PIL import Image as PILImage

'''
****************** DEFINING FUNCTIONS ******************
FUNCTION DESCRIPTOR TEMPLATE
PURPOSE/DESCRIPTION: 
INPUT: 
OUTPUT: 
'''

def colLettersToNumber(letters):
    '''
    PURPOSE/DESCRIPTION: Columns can be assigned numbers. For example, Col A in Excel is the same as Col 0 (by default). 
    This function takes in the letter of a column (e.g., Col A) and converts it to its numerical equivalent. 
    This makes it significantly easier to perform cell offset operations (e.g., 2 cells left of D4 is B4). Importantly, 
    this function defines Col A as Col 1, Col B as Col 2, Col C as Col 3, etc. 

    INPUT: 
    letters (String): The column letter to be converted into numerical form (with Col A as Col 1).  

    OUTPUT: 
    running_total (int): The respective column number from the input column letter (with Col A as Col 1). 
    '''
    # If the input 'letters' is in the alphabet string below, return the index plus 1 (so that it starts at Col A as Col 1).
    letter_list = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    if letters in letter_list:
        return letter_list.index(letters) + 1
    
    # Otherwise, if it is multi-lettered (e.g., 'BG' or 'ABN'), calculate the conversion with base 26 number system. 
    else:
        running_total = 0
        place = len(letters) - 1
        for l in letters:
            running_total = running_total + (26 ** place) * (letter_list.index(l) + 1)
            place = place - 1

        return running_total

def colNumberToLetters(n):
    '''
    PURPOSE/DESCRIPTION: Columns can be assigned numbers. For example, Col A in Excel is the same as Col 0 (by default). 
    This function takes in the number of a column and converts it into its letter equivalent. This makes it significantly 
    easier to perform cell offset operations using numbers (e.g., 2 cells left of D4 is B4) and to obtain column letters 
    again. Importantly, this function defines Col 1 as Col A, Col 2 as Col B, Col 3 as Col C, etc. 

    INPUT: 
    n (int): The column number to be converted into letter form (with Col 1 as Col A).  

    OUTPUT: 
    column_letter (String): The respective column letter from the input column number (with Col 1 as Col A). 
    '''
    # Converts the provided number into unicode letters using a base 26 numerical system. 
    column_letter = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        column_letter = chr(65 + remainder) + column_letter

    return column_letter

def relativeToCell(cell_string, translations):
    '''
    PURPOSE/DESCRIPTION: It is useful to reference cells not absolutely, but in reference to other specified cells. 
    This function will take the address (column letter and row number, e.g., DE54) of a specified cell and a list of 
    translations with its origin at the specified cell, and run through the list of translations until it arrives
    at the desired cell. For example, calling this function on Cell H5 with the following translations: 
        - Up, 1
        - Right, 2
        - Left, 3
        - Down, 2
    Will return the cell address of Cell G6. 
    Or
    relativeToCell(ws, 'B2', [('right', 4), ('down', 2)]) should return 'F4'

    INPUT: 
    cell_string (String): The cell address for the translations to be conducted in reference to.   
    translations (List): A list of tuples with containing the translations with respect to the cell in cell_string that 
    will result in the desired cell. In other words, relative to the cell at cell_string, these translations will result 
    in this desired cell. The format of translations is as follows: 
    [(direction, units), (direction, units), ... , (direction, units), (direction, units)]
    Where direction is 'right', 'left', 'up', or 'down'
    and units is any integer which may be negative, positive, or 0. 

    OUTPUT: 
    desired_cell (String): The cell address one arrives at when following the list of translations, with respect
    to the input cell address. 
    '''
    # Break down the input cell string. 
    current_letters, current_numbers = splitCellCoord(cell_string)
    current_numbers = int(current_numbers)

    # Update the letter and number portions of the current cell based on the provided translations. 
    for movement in translations:
        dir = movement[0]
        units = movement[1]
        if dir == 'right':
            current_letters = colNumberToLetters(colLettersToNumber(current_letters) + units)
        elif dir == 'left':
            current_letters = colNumberToLetters(colLettersToNumber(current_letters) - units)
        elif dir == 'down':
            current_numbers = current_numbers + units
        elif dir == 'up':
            current_numbers = current_numbers - units

    # Raise some exceptions to invalid translations; the function does not have a default behavior when bad translation inputs are put in.
    if current_letters == '':
        raise Exception('Target column is out of range')   
    if current_numbers < 0:
        raise Exception("Target row is out of range")
    
    # Combine the letter and number portions to obtain the address of the new cell location that is relative to the reference cell via the translation list and return. 
    desired_cell = current_letters + str(current_numbers)
    return desired_cell

def splitCellCoord(coord):
    '''
    PURPOSE/DESCRIPTION: It is convenient to split a cell's address, such as A4, B9, AB98, and so on. Because 
    the letters representing the column could be more than one letter and there is no delimiter between letter
    and number, a more creative approach must be taken to split the letter and number apart, hence this function. 

    INPUT: 
    coord (String): The string representing a cell's address (e.g., "F4", "BV88", etc.).

    OUTPUT: 
    split_tuple (Tuple): A tuple consisting of the letter substring of coord as the 0th element and the numeric
    substring as the 1st element. 
    '''
    # Determine where the split from letters to numbers is in the provided cell address, index-wise. 
    i_cell = len(coord) - 1
    while coord[i_cell].isnumeric():
        i_cell = i_cell - 1

    # Split into the respective letters and numbers substrings. 
    cell_letters = coord[:i_cell + 1]
    cell_numbers = coord[i_cell + 1:]

    # Throw these two substrings into a tuple and return. 
    split_tuple = (cell_letters, cell_numbers)
    return split_tuple

def fillCellColors(worksheet, range_start, range_end, color, f_type):
    '''
    PURPOSE/DESCRIPTION: Fills all of the cells within a rectuangular range to be a specified color. 

    INPUT: 
    worksheet (Worksheet): The desired worksheet that this function is to be used on. 
    range_start (String): The top-left cell address (e.g., "B2") of the desired range to color fill. 
    range_end (String): The bottom-right cell address (e.g., "G8") of the desired range to color fill. 
    color (String): The fill color as hexadecimal (e.g. 'A6C9EC', 'DAE9F8', '83CCEB', 'C6C9EC'). 
    f_type (String): The type of fill. Usually 'solid'. 

    OUTPUT: 
    N/A. 
    '''
    # Iterate through all cells in the provided range and fill. 
    fill = PatternFill(start_color = color, end_color = color, fill_type = f_type)
    for row in worksheet[range_start + ":" + range_end]:
        for cell in row:
            cell.fill = fill

def setColumnWidths(worksheet, desired_width):
    '''
    PURPOSE/DESCRIPTION: Sets Column A through Column Z to be a specified width. 

    INPUT: 
    worksheet (Worksheet): The worksheet where this function will be applied. 
    desired_width (int/double/float): The desired width of Column A through Column Z. 

    OUTPUT: 
    N/A. 
    '''
    # Iterate through all the columns from A to Z and change the column width. 
    cols = list('ABCDEFGHIJKLMNOPQRSTUVWXYZ')
    for col in cols:
        worksheet.column_dimensions[col].width = desired_width

def createThickOutsideBorders(worksheet, range_start, range_end):
    '''
    PURPOSE/DESCRIPTION: To create thick outside borders around the specified rectangular range. 

    INPUT: 
    worksheet (Worksheet): The worksheet where this function will apply on. 
    range_start (String): The top-left of the range to be thick bordered. 
    range_end (String): The bottom-right of the range to be thick bordered. 

    OUTPUT: 
    N/A.
    '''
    # Split the cell addresses of the range bounds for ease of use. 
    start_letters, start_numbers = splitCellCoord(range_start)
    end_letters, end_numbers = splitCellCoord(range_end)

    # For every row in the specified range...
    for row in worksheet[range_start + ":" + range_end]:
        # For every cell in said row...
        for cell in row: 

            # Determine the cell's address and split into letters and numbers. 
            coord = cell.coordinate
            cell_letters, cell_numbers = splitCellCoord(coord)

            # Determine if the cell is a corner cell of the range and if so, create that corner's thick outside border. 
            if cell_letters == start_letters and cell_numbers == start_numbers:
                worksheet[coord].border = Border(left=Side(style='thick'), top=Side(style='thick'))
            elif cell_letters == start_letters and cell_numbers == end_numbers:
                worksheet[coord].border = Border(left=Side(style='thick'), bottom=Side(style='thick'))
            elif cell_letters == end_letters and cell_numbers == start_numbers:
                worksheet[coord].border = Border(right=Side(style='thick'), top=Side(style='thick'))
            elif cell_letters == end_letters and cell_numbers == end_numbers:
                worksheet[coord].border = Border(right=Side(style='thick'), bottom=Side(style='thick'))

            # Determine if the cell is an edge cell of the range and if so, create that cell edge's thick outside border. 
            elif cell_letters == start_letters:
                worksheet[coord].border = Border(left=Side(style='thick'))
            elif cell_letters == end_letters:
                worksheet[coord].border = Border(right=Side(style='thick'))
            elif cell_numbers == start_numbers:
                worksheet[coord].border = Border(top=Side(style='thick'))
            elif cell_numbers == end_numbers:
                worksheet[coord].border = Border(bottom=Side(style='thick'))
            
def createIntersectionBorders(worksheet, main_coords):
    '''
    PURPOSE/DESCRIPTION: To create the linework representing the intersection itself. 

    INPUT: 
    worksheet (Worksheet): The worksheet in which this function is to be applied on. 
    main_coords (List of Strings): A list of cell addresses, with the 0th element being the cell address 
    of the top-left corner of the "main" figure area (everything but the header) and the 1st element
    being the cell address of the bottom-right corner. 

    OUTPUT: 

    N/A. 
    '''
    # Split the cell addresses of the range bounds for ease of use. 
    top_left_letter, top_left_number = splitCellCoord(main_coords[0])
    bottom_right_letter, bottom_right_number = splitCellCoord(main_coords[1])

    # Convert numbers in String format to actual numbers.
    top_left_number = int(top_left_number)
    bottom_right_number = int(bottom_right_number)

    # If the main display width consists of an even number of columns, calculate the midpoint this way. Otherwise if it is an odd number, calculate it the other way (they are both currently the same). 
    if main_display_width % 2 == 0:
        hori_midpoint = math.floor((colLettersToNumber(bottom_right_letter) - colLettersToNumber(top_left_letter))/2 + colLettersToNumber(top_left_letter))
    else: 
        hori_midpoint = math.floor((colLettersToNumber(bottom_right_letter) - colLettersToNumber(top_left_letter))/2 + colLettersToNumber(top_left_letter))

    # If the main display height consists of an even number of rows, calculate the midpoint this way. Otherwise if it is an odd number, calculate it the other way
    if main_display_height % 2 == 0:
        vert_midpoint = math.floor((bottom_right_number - top_left_number)/2 + top_left_number)
    else: 
        vert_midpoint = math.floor((colLettersToNumber(bottom_right_letter) - colLettersToNumber(top_left_letter))/2 + colLettersToNumber(top_left_letter))

    # Determine the midpoint cell. 
    hori_midpoint_cell = colNumberToLetters(hori_midpoint) + str(top_left_number + 1)
    vert_midpoint_cell = colNumberToLetters(colLettersToNumber(top_left_letter) + 1) + str(vert_midpoint)
    
    # Determine the cells where the horizontal and vertical interseciton lines will stop at. 
    vert_end = relativeToCell(hori_midpoint_cell, [('down', main_display_height-3)])
    hori_end = relativeToCell(vert_midpoint_cell, [('right', main_display_width-3)])

    # Keep drawing the vertical intersection line until we reach the specified end (vert_end). 
    current_vert = hori_midpoint_cell
    reachedVertEnd = False
    while not reachedVertEnd:
        if current_vert == vert_end:
            reachedVertEnd = True
        worksheet[current_vert].border = Border(right=Side(style='thin'))
        current_vert = relativeToCell(current_vert, [('down', 1)])

    # Keep drawing the horizontal intersection line until we reach the specified end (hori_end). 
    current_hori = vert_midpoint_cell
    reachedHoriEnd = False
    while not reachedHoriEnd:
        if current_hori == hori_end:
            reachedHoriEnd = True
        worksheet[current_hori].border = Border(bottom=Side(style='thin'))
        current_hori = relativeToCell(current_hori, [('right', 1)])

    # Perform special border instructions for the middle cell specifically (because otherwise, it will be missing a border side). 
    middle_cell = splitCellCoord(hori_midpoint_cell)[0] + str(splitCellCoord(vert_midpoint_cell)[1])
    ws[middle_cell].border = Border(bottom=Side(style='thin'), right=Side(style='thin'))    

def generateFigure(ws, df, df_row, origin):
    '''
    PURPOSE/DESCRIPTION: Generates one individual figure. 

    INPUT:
    ws (Worksheet): The worksheet the figure is to be generated on. 
    df (Dataframe): The Pandas dataframe to be used to pull information from. 
    df_row (Dataframe row): A row in the Pandas dataframe corresponding to the intersection to which a figure should be generated. 
    origin (String): The cell address defining the "origin" of the figure that most figure creation processes refer off of. Top-left cell of figure. 

    OUTPUT: 
    N/A.
    '''
    # Make the lists. 
    header_coords = [None, None]
    main_coords = [None, None]

    # If there is a header desired, create the header and appropriately set the range bounds for the header and the range bounds for the main display area. 
    if header:
        header_coords = [origin, relativeToCell(origin, [('right', main_display_width-1), ('down', header_height-1)])]
        main_coords = [relativeToCell(origin, [('down', 2)]), relativeToCell(origin, [('right', main_display_width-1), ('down', main_display_height + header_height-1)])]
        fillCellColors(ws, header_coords[0], header_coords[1], header_color, 'solid')
        createThickOutsideBorders(ws, header_coords[0], header_coords[1])
        ws.merge_cells(header_coords[0] + ":" + header_coords[1])
        
    # Otherwise, since there is no header desired, just set the range bounds for the main display area appropriately. 
    else:
        main_coords = [origin, relativeToCell(origin, [('right', main_display_width-1), ('down', main_display_height-1)])]

    # If there is a border of different colored cells surrounding the main display area desired, fill in these cells with the specified border color. 
    if main_border:
        fillCellColors(ws, main_coords[0], main_coords[1], main_border_color, 'solid')
        fillCellColors(ws, relativeToCell(main_coords[0], [('right', 1), ('down', 1)]), relativeToCell(main_coords[1], [('left', 1), ('up', 1)]), main_bkgd_color, 'solid')
    
    # Otherwise, just fill in these cells with whatever the main display area color is.  
    else:
        fillCellColors(ws, main_coords[0], main_coords[1], main_bkgd_color, 'solid')

    # *** THIS IS A DEBUG CODE BLOCK ***
    # print('origin: ' + str(origin))
    # print('header: ' + str(header_coords))
    # print('main: ' + str(main_coords))
    # *** THIS IS A DEBUG CODE BLOCK ***

    # If an intersection number box is desired, then...
    if int_num_box: 

        # If there is also a main border, merge these following cells for the intersection number box. 
        if not main_border:
            box_topleft = main_coords[0]
            box_bottomright = relativeToCell(box_topleft, [('down', 1), ('right', 1)])
            ws.merge_cells(box_topleft + ":" + box_bottomright)
        # Otherwise, merge these following cells for the intersection number box. 
        else:
            box_topleft = relativeToCell(main_coords[0], [('down', 1), ('right', 1)])
            box_bottomright = relativeToCell(box_topleft, [('down', 1), ('right', 1)])
            ws.merge_cells(box_topleft + ":" + box_bottomright)
        # Format the intersection number box and assign it the proper intersection number. 
        fillCellColors(ws, box_topleft, box_bottomright, int_num_box_color, 'solid')
        int_num = df.loc[df_row, 'Int. ID 1']
        ws[box_topleft].value = int_num
        ws[box_topleft].alignment = Alignment(horizontal='center', vertical='center')

    # If cardinal directions are desired...
    if cardinal_dirs:
        # Split the cell addresses to make it easier to work with. 
        top_left_letter, top_left_number = splitCellCoord(main_coords[0])
        bottom_right_letter, bottom_right_number = splitCellCoord(main_coords[1])

        # Convert String numbers to actual numbers. 
        top_left_number = int(top_left_number)
        bottom_right_number = int(bottom_right_number)

        # If the width of the main display area is even, then merge the following cells to obtain cell_north and cell_south, to keep the 'N' and 'S' centered on the figure. 
        if main_display_width % 2 == 0:
            hori_midpoint = math.floor((colLettersToNumber(bottom_right_letter) - colLettersToNumber(top_left_letter))/2 + colLettersToNumber(top_left_letter))
            north_merge_range_str = colNumberToLetters(hori_midpoint) + str(top_left_number) + ":" + colNumberToLetters(hori_midpoint + 1) + str(top_left_number)
            ws.merge_cells(north_merge_range_str)
            north_merge_range = north_merge_range_str.split(':')
            cell_north = ws[north_merge_range[0]]

            south_merge_range_str = relativeToCell(north_merge_range[0], [('down', main_display_height-1)]) + ":" + relativeToCell(north_merge_range[0], [('down', main_display_height-1), ('right', 1)])
            ws.merge_cells(south_merge_range_str)
            south_merge_range = south_merge_range_str.split(":")
            cell_south = ws[south_merge_range[0]]

        # Otherwise, if the width of the main display is odd, then no merging is necessary (because one cell will be the true "center"). 
        else: 
            hori_midpoint = math.floor((colLettersToNumber(bottom_right_letter) - colLettersToNumber(top_left_letter))/2 + colLettersToNumber(top_left_letter))
            cell_north = ws[colNumberToLetters(hori_midpoint) + str(top_left_number)]
            cell_south = ws[relativeToCell(colNumberToLetters(hori_midpoint) + str(top_left_number), [('down', main_display_height-1)])]

        # If the height of the main display area is even, then merge the following cells to obtain cell_east and cell_west, to keep the 'E' and 'W' centered on the figure. 
        if main_display_height % 2 == 0:
            vert_midpoint = math.floor((bottom_right_number - top_left_number)/2 + top_left_number)
            west_merge_range_str = top_left_letter + str(vert_midpoint) + ":" + top_left_letter + str(vert_midpoint + 1)
            ws.merge_cells(west_merge_range_str)
            west_merge_range = west_merge_range_str.split(':')
            cell_west = ws[west_merge_range[0]]

            east_merge_range_str = relativeToCell(west_merge_range[0], [('right', main_display_width-1)]) + ":" + relativeToCell(west_merge_range[0], [('right', main_display_width-1), ('down', 1)])
            ws.merge_cells(east_merge_range_str)
            east_merge_range = east_merge_range_str.split(":")
            cell_east = ws[east_merge_range[0]]

        # Otherwise, if the height of the main display is odd, then no merging is necessary (because one cell will be the true "center"). 
        else: 
            vert_midpoint = math.floor((colLettersToNumber(bottom_right_letter) - colLettersToNumber(top_left_letter))/2 + colLettersToNumber(top_left_letter))
            cell_west = ws[top_left_letter + str(vert_midpoint)]
            cell_east = ws[relativeToCell(top_left_letter + str(vert_midpoint), [('down', main_display_height-1)])]

        # Assign the value 'N' to the north cardinal direction cell and format. 
        cell_north.value = 'N'
        cell_north.alignment = Alignment(horizontal='center', vertical='center')
        cell_north.font = Font(bold=True)
        
        # Assign the value 'S' to the south cardinal direction cell and format. 
        cell_south.value = 'S'
        cell_south.alignment = Alignment(horizontal='center', vertical='center')
        cell_south.font = Font(bold=True)

        # Assign the value 'W' to the west cardinal direction cell and format. 
        cell_west.value = 'W'
        cell_west.alignment = Alignment(horizontal='center', vertical='center')
        cell_west.font = Font(bold=True)
        
        # Assign the value 'E' to the east cardinal direction cell and format. 
        cell_east.value = 'E'
        cell_east.alignment = Alignment(horizontal='center', vertical='center')
        cell_east.font = Font(bold=True)

    # Create the thick outside borders for the figure, create the intersection border linework, and create the intersection number box, if applicable. 
    createThickOutsideBorders(ws, main_coords[0], main_coords[1])
    createIntersectionBorders(ws, main_coords)
    if int_num_box:
        createThickOutsideBorders(ws, box_topleft, box_bottomright)


    # THOMAS SELF-COMMENT
    # THIS CODE BELOW WILL LIKELY NEED CODE TO TACKLE THE INSTANCE WHERE THERE ARE NO SIGNS AT ALL
    # PROBABLY A TRY-CATCH OR SOMETHING ELSE TO DETECT IF THE SIGN IS A NAN
    # 

    # Insert the northbound sign into the figure. 
    nb_sign = df.loc[df_row, '@image_NB_sign'][:-3] + 'png'
    nb_sign_coord = relativeToCell(cell_east.coordinate, [('down', 2), ('left', 2)])
    insertImageWithOffset(ws, nb_sign_coord, img_dir_path, nb_sign, 0, 30, 30, 0.00, 0.00)

    # Insert the southbound sign into the figure. 
    sb_sign = df.loc[df_row, '@image_SB_sign'][:-3] + 'png'
    sb_sign_coord = relativeToCell(cell_west.coordinate, [('up', 1), ('right', 1)])
    insertImageWithOffset(ws, sb_sign_coord, img_dir_path, sb_sign, 180, 30, 30, 0.60, -0.50)

    # Insert the westbound sign into the figure. 
    wb_sign = df.loc[df_row, '@image_WB_sign'][:-3] + 'png'
    wb_sign_coord = relativeToCell(cell_north.coordinate, [('down', 1), ('right', 1)])
    insertImageWithOffset(ws, wb_sign_coord, img_dir_path, wb_sign, 90, 30, 30, 1.00, 0.60)

    # Insert the eastbound sign into the figure. 
    eb_sign = df.loc[df_row, '@image_EB_sign'][:-3] + 'png'
    eb_sign_coord = relativeToCell(cell_south.coordinate, [('up', 2), ('left', 1)])
    insertImageWithOffset(ws, eb_sign_coord, img_dir_path, eb_sign, 270, 30, 30, -0.50, 0.00)
    
    # Insert a signal image if the intersection row in the dataframe has one, if not, do nothing. 
    try: 
        signal = df.loc[df_row, '@image_light'][:-3] + 'png'
        signal_coord = relativeToCell(cell_north.coordinate, [('down', int(main_display_height/2) - 1)])
        insertImageWithOffset(ws, signal_coord, img_dir_path, signal, 0, 42, 42, -0.10, -0.10)
    except TypeError:
        dummy = 'do nothing'

    return None

def importVolumes(ws, df, df_row, origin, travel_dir):
    '''
    PURPOSE/DESCRIPTION: Import the volumes into a figure based on intersection and travel direction. Note that this currently assumes that the turn arrow imagery are image files with
    an extension of three letters (e.g., .png, .jpg, etc.). This code will not work if the file extension is not exactly three letters long (e.g., potentially .jpeg). 

    INPUT: 
    ws (Worksheet): The worksheet on which to import the volumes onto. 
    df (Dataframe): The Pandas dataframe on which the data is being pulled from. 
    df_row (Dataframe row): A row in the Pandas dataframe corresponding to the intersection to which a figure should be generated. 
    origin (String): The cell address defining the "origin" of the figure that most figure creation processes refer off of. Top-left cell of figure. 
    travel_dir (String): One of the following: "NB", "SB", "EB", or "WB". Denotes which travel direction volumes are to be imported. More can be added into the code. 

    OUTPUT: 
    N/A.
    '''
    # Shorten the variable names. 
    height = main_display_height
    width = main_display_width

    # Create the lane names by appending the direction of travel (e.g., 'NB' or 'SB') with the lane number (a number from 1 through 6, inclusive).
    lanes = []
    for i in range(1,7):
        lanes.append(travel_dir + str(i))

    # Determine where to start importing the volumes in relation to the figure origin cell (top-left cell) based on direction of travel by choosing the appropriate translation list. 
    if travel_dir == 'EB':
        translation_list = [('down', int(height/2 + 3)), ('right', int(width/2 - 4))]
    elif travel_dir == 'NB':
        translation_list = [('down', int(height/2 + 3)), ('right', int(width/2 + 4))]
    elif travel_dir == 'WB':
        translation_list = [('down', int(height/2 - 3)), ('right', int(width/2 + 3))]
    elif travel_dir == 'SB':
        translation_list = [('down', int(height/2 - 4)), ('right', int(width/2 - 4))]

    # Modify the translation list to account for the header, if it exists. 
    if header:
        translation_list.append(('down', header_height))

    # With the chosen translation list, determine where to start importing the volumes in based on direction of travel. 
    lane_coord = relativeToCell(origin, translation_list)

    # For each lane of the six total lanes for that direction...
    for lane_num in lanes: 

        # *** THIS IS A DEBUG CODE BLOCK ***
        # print(df.loc[df_row, 'Scenario'] + ' ' + lane_num)
        # print(df.loc[df_row, lane_num])
        # *** THIS IS A DEBUG CODE BLOCK ***

        # Create the strings that will be used in searching the dataframe for the turn arrows needed and put into a list. 
        lane_arrows = []
        for i in range(0,3):
            lane_arrows.append('@image_' + lane_num + '.' + str(i))

        # Locate the volume belonging to lane_num and write it into the worksheet at lane_coord. 
        ws[lane_coord] = df.loc[df_row, lane_num]
        
        # If the direction of travel is eastbound, determine and insert the appropriate eastbound turn arrows, as well as format. 
        if travel_dir == 'EB':
            ws[lane_coord].alignment = Alignment(horizontal='right', vertical='center')
            lane_arrow_coord = relativeToCell(lane_coord, [('right', 1)])
            for arrow in lane_arrows:
                try: 
                    img_file_name = df.loc[df_row, arrow]
                    img_file_name = img_file_name[:-3] + 'png'
                    insertImageWithOffset(ws, lane_arrow_coord, img_dir_path, img_file_name, 0, 30, 30, 0.20, -0.30)
                except IndexError:
                    dummy = 'do nothing'
                except TypeError:
                    dummy = 'do nothing'

            # Move to where the next eastbound lane volume should be imported. 
            lane_coord = relativeToCell(lane_coord, [('down', 1)])  

        # If the direction of travel is northbound, determine and insert the appropriate northbound turn arrows, as well as format. 
        elif travel_dir == 'NB':
            cells_to_merge_nb = lane_coord + ":" + relativeToCell(lane_coord, [('down', int(height/2)-5)])
            ws.merge_cells(cells_to_merge_nb)
            ws[lane_coord].alignment = Alignment(textRotation=90, horizontal='center', vertical='top')
            lane_arrow_coord = relativeToCell(lane_coord, [('up', 1)])
            for arrow in lane_arrows:
                try: 
                    img_file_name = df.loc[df_row, arrow]
                    img_file_name = img_file_name[:-3] + 'png'
                    insertImageWithOffset(ws, lane_arrow_coord, img_dir_path, img_file_name, 90, 30, 30, -0.45, -0.65)
                except IndexError:
                    dummy = 'do nothing'
                except TypeError:
                    dummy = 'do nothing'

            # Move to where the next northbound lane volume should be imported. 
            lane_coord = relativeToCell(lane_coord, [('right', 1)])

        # If the direction of travel is westbound, determine and insert the appropriate westbound turn arrows, as well as format. 
        elif travel_dir == 'WB':
            ws[lane_coord].alignment = Alignment(horizontal='left', vertical='center')
            lane_arrow_coord = relativeToCell(lane_coord, [('left', 1)])
            for arrow in lane_arrows:
                try: 
                    img_file_name = df.loc[df_row, arrow]
                    img_file_name = img_file_name[:-3] + 'png'
                    insertImageWithOffset(ws, lane_arrow_coord, img_dir_path, img_file_name, 180, 30, 30, -0.80, -0.15)
                except IndexError:
                    dummy = 'do nothing'
                except TypeError:
                    dummy = 'do nothing'

            # Move to where the next westbound lane volume should be imported. 
            lane_coord = relativeToCell(lane_coord, [('up', 1)])

        # If the direction of travel is southbound, determine and insert the appropriate southbound turn arrows, as well as format. Southbound has a special case of doing ws[lane_coord] = df.loc[df_row, lane_num]
        # because of the cell addresses of merged cells. The original ws[lane_coord] = df.loc[df_row, lane_num] assignment to the worksheet is wiped away when the cell was merged, I believe. 
        elif travel_dir == 'SB':
            cells_to_merge_sb = relativeToCell(lane_coord, [('up', int(height/2)-5)]) + ":" + lane_coord
            ws.merge_cells(cells_to_merge_sb)
            lane_coord_sb_merged = relativeToCell(lane_coord, [('up', int(height/2)-5)])
            ws[lane_coord_sb_merged].alignment = Alignment(textRotation=90, horizontal='center', vertical='bottom')
            ws[lane_coord_sb_merged] = df.loc[df_row, lane_num]
            lane_arrow_coord = relativeToCell(lane_coord, [('down', 1)])
            for arrow in lane_arrows:
                try: 
                    img_file_name = df.loc[df_row, arrow]
                    img_file_name = img_file_name[:-3] + 'png'
                    insertImageWithOffset(ws, lane_arrow_coord, img_dir_path, img_file_name, 270, 30, 30, -0.25, 0.20)
                except IndexError:
                    dummy = 'do nothing'
                except TypeError:
                    dummy = 'do nothing'

            # Move to where the next southbound lane volume should be imported. 
            lane_coord = relativeToCell(lane_coord, [('left', 1)])

def insertImage(ws, coord, img_path, img_name, rotation, img_height, img_width):
    '''
    PURPOSE/DESCRIPTION: Inserts an image into the worksheet without the option for offsets (i.e., the top-left corner of the image
    will always be aligned with the top-left corner of a specified cell). Mostly unused at this point, because of the more versatile 
    insertImageWithOffset function (where you can just specify no offset), but keeping this function here anyway. 

    INPUT: 
    ws (Worksheet): The worksheet on which the image is to be inserted. 
    coord (String): The cell address (e.g., "B2") on which the image is to be anchored/inserted in reference to. The top-left corner of the 
    image will be aligned with this cell. 
    img_path (String): The path to the directory (i.e., do not include image file name in the path) in which all of the images are stored. 
    img_name (String): The file name of the image to be inserted, including extensions. 
    rotation (int/double/float): The amount (in degrees, counterclockwise) the inserted image should be rotated. This rotated image is stored 
    in a separate directory called temp in the img_path. The temp directory should pre-exist before this function executes (this function 
    will not create the temp directory).
    img_height (int/double/float): The desired height of the image (in what I assume to be pixels, but am not sure). 
    img_width (int/double/float): The desired width of the image (in what I assume to be pixels, but am not sure). 

    OUTPUT: 
    N/A. 
    '''
    # Rotate the image as desired, save this new rotated image to a new file in a pre-existing temp directory. 
    pil_img = PILImage.open(img_path + img_name)
    rotated_img = pil_img.rotate(rotation, expand=True)
    temp_img_path = img_path + "\\temp\\" + str(rotation) + "_" + img_name[1:]
    rotated_img.save(temp_img_path)

    # *** THIS IS A DEBUG CODE BLOCK ***
    # print("Original Image Path")
    # print(img_path + img_name)
    # print("Temporary Image Path")
    # print(temp_img_path)
    # *** THIS IS A DEBUG CODE BLOCK ***

    # Create image object, alter its dimensions, and insert into figure. 
    img = Image(temp_img_path)
    img.width = img_width
    img.height = img_height  
    ws.add_image(img, coord)
    return

def insertImageWithOffset(ws, coord, img_path, img_name, rotation, img_height, img_width, x_offset, y_offset):
    '''
    PURPOSE/DESCRIPTION: Inserts an image into the worksheet with the option for offsets as a percentage of column width. 
    If column width or row height is ever adjusted from what it currently is in the code, then x_offset and y_offset will 
    still work but will likely require trial and error to get your desired offset. 

    INPUT: 
    ws (Worksheet): The worksheet on which the image is to be inserted. 
    coord (String): The cell address (e.g., "B2") on which the image is to be anchored/inserted in reference to. The top-left corner of the 
    image will be aligned with this cell, and any offsets will deviate the inserted image from this cell's top-left corner. 
    img_path (String): The path to the directory (i.e., do not include image file name in the path) in which all of the images are stored. 
    img_name (String): The file name of the image to be inserted, including extensions. 
    rotation (int/double/float): The amount (in degrees, counterclockwise) the inserted image should be rotated. This rotated image is stored 
    in a separate directory called temp in the img_path. The temp directory should pre-exist before this function executes (this function 
    will not create the temp directory).
    img_height (int/double/float): The desired height of the image (in what I assume to be pixels, but am not sure). 
    img_width (int/double/float): The desired width of the image (in what I assume to be pixels, but am not sure). 
    x_offset (int/double/float): The desired horizontal offset relative to the cell located at coord, in terms of coord's width (e.g., 
    having x_offset be 0.50 would offset the inserted image half of coord's width to the right).
    y_offset (int/double/float): The desired vertical offset relative to the cell located at coord, in terms of coord's height (e.g., 
    having y_offset be 0.50 would offset the inserted image half of coord's height downwards).

    OUTPUT: 
    N/A. 
    '''
    # Rotate the image as desired, save this new rotated image to a new file in a pre-existing temp directory; create the image object. 
    pil_img = PILImage.open(img_path + img_name)
    rotated_img = pil_img.rotate(rotation, expand=True)
    temp_img_path = img_path + "\\temp\\" + str(rotation) + "_" + img_name[1:]
    rotated_img.save(temp_img_path)
    img = Image(temp_img_path)

    # *** THIS IS A DEBUG CODE BLOCK ***
    # print("Original Image Path")
    # print(img_path + img_name)
    # print("Temporary Image Path")
    # print(temp_img_path)
    # *** THIS IS A DEBUG CODE BLOCK ***

    # Perform these steps. 
    p2e = pixels_to_EMU
    size = XDRPositiveSize2D(p2e(img_width), p2e(img_height))
    c2e = cm_to_EMU

    # Define anonymous functions based on default Excel cell heights and widths, except that I modified the cell width lambda function by multiplying 
    # by 2/8.09 (current cell width / default cell width for our purposes. 
    cellh = lambda x: c2e((x * 49.77)/99)
    cellw = lambda x: c2e((x * (18.65-1.71))/10 * (2/8.09))

    # Split img_anchor cell address and determine convert offsets from a percentage of cell dimensions to EMU (I presume). 
    img_anchor = splitCellCoord(coord)
    column = colLettersToNumber(img_anchor[0]) - 1 # Minus 1 to account for our system doing Col A = Col 1 and Row 1 = Row 1, while AnchorMarker does Col A = Col 0 and Row 1 = Row 0
    row = int(img_anchor[1]) - 1
    coloffset = cellw(x_offset)
    rowoffset = cellh(y_offset)

    # Insert the image. 
    marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    ws.add_image(img) 
    return None

def populateFigure(ws, df, df_row, origin):
    '''
    PURPOSE/DESCRIPTION: Populates the figures with volumes, names, etc. 

    INPUT: 
    ws (Worksheet): The worksheet in which the figure is located to be populated. 
    df (Dataframe): The Pandas dataframe to be used to pull information from. 
    df_row (Dataframe row): A row in the Pandas dataframe corresponding to the intersection to which a figure should be generated. 
    origin (String): The cell address defining the "origin" of the figure that most figure creation processes refer off of. Top-left cell of figure. 

    OUTPUT: 
    N/A.
    '''
    # Shorten the variable names. 
    height = main_display_height
    width = main_display_width

    # Determine what is the current scenario and write it to the worksheet next to the figure. This is more of a debugging tool. 
    scenario = df.loc[df_row, 'Scenario']
    scenario_cell = relativeToCell(origin, [('right', width + 1)])
    ws[scenario_cell] = scenario
    
    # [STILL NEEDS TO ACCOMMODATE OVERFLOW TEXT, ALSO DOUBLE CHECK HOW THESE KEYS COME (WHETHER THERE IS A SPACE OR NOT)]

    # Insert the eastbound road name.  
    eb_roadname = df.loc[df_row, 'EB Road Name ']
    eb_translate_list = [('down', int(height/2)), ('right', 1)]
    if header:
        eb_translate_list.append(('down', header_height))
    eb_roadname_cell = relativeToCell(origin, eb_translate_list)
    ws[eb_roadname_cell] = eb_roadname

    # Insert the westbound road name.  
    wb_roadname = df.loc[df_row, 'WB Road Name']
    wb_translate_list = [('down', int(height/2 - 1)), ('right', width - 2)]
    if header:
        wb_translate_list.append(('down', header_height))
    wb_roadname_cell = relativeToCell(origin, wb_translate_list)
    ws[wb_roadname_cell] = wb_roadname
    ws[wb_roadname_cell].alignment = Alignment(horizontal='right')

    # Insert the southbound road name.  
    sb_roadname = df.loc[df_row, 'SB Road Name']
    sb_translate_list_topleft = [('right', int(width/2)-1), ('down', 1)]
    sb_translate_list_bottomright = [('right', int(width/2)-1), ('down', int(height/2-1))]
    if header:
        sb_translate_list_topleft.append(('down', header_height))
        sb_translate_list_bottomright.append(('down', header_height))
    sb_roadname_topleft_cell = relativeToCell(origin, sb_translate_list_topleft)
    sb_roadname_bottomright_cell = relativeToCell(origin, sb_translate_list_bottomright)
    sb_roadname_cellstomerge = sb_roadname_topleft_cell + ":" + sb_roadname_bottomright_cell
    ws.merge_cells(sb_roadname_cellstomerge)
    ws[sb_roadname_topleft_cell] = sb_roadname
    ws[sb_roadname_topleft_cell].alignment = Alignment(textRotation=90, vertical='top')

    # Insert the northbound road name.  
    nb_roadname = df.loc[df_row, 'NB Road Name ']
    nb_translate_list_topleft = [('right', int(width/2)), ('down', int(height/2))]
    nb_translate_list_bottomright = [('right', int(width/2)), ('down', int(height)-2)]
    if header:
        nb_translate_list_topleft.append(('down', header_height))
        nb_translate_list_bottomright.append(('down', header_height))
    nb_roadname_topleft_cell = relativeToCell(origin, nb_translate_list_topleft)
    nb_roadname_bottomright_cell = relativeToCell(origin, nb_translate_list_bottomright)
    nb_roadname_cellstomerge = nb_roadname_topleft_cell + ":" + nb_roadname_bottomright_cell
    ws.merge_cells(nb_roadname_cellstomerge)
    ws[nb_roadname_topleft_cell] = nb_roadname
    ws[nb_roadname_topleft_cell].alignment = Alignment(textRotation=90, vertical='bottom')

    # Import the volumes for all four directions. 
    importVolumes(ws, df, df_row, origin, 'EB')
    importVolumes(ws, df, df_row, origin, 'NB')
    importVolumes(ws, df, df_row, origin, 'WB')
    importVolumes(ws, df, df_row, origin, 'SB')

    # NEED TO INSERT LOGIC FOR NO STREET NAMES, NAN IS COMING OUT AS FLOAT
    if header:
        header_str = ''

        if type(nb_roadname) == float and math.isnan(nb_roadname): 
            nb_roadname = ""
        if type(sb_roadname) == float and math.isnan(sb_roadname):
            sb_roadname = ""
        if type(eb_roadname) == float and math.isnan(eb_roadname):
            eb_roadname = ""
        if type(wb_roadname) == float and math.isnan(wb_roadname):
            wb_roadname = ""

        # If both road names the same and not empty, add the road name and an '&' to the header string so far. 
        if nb_roadname == sb_roadname and len(nb_roadname) > 0: 
            header_str = header_str + nb_roadname + " & "
        # If nb_roadname is empty while sb_roadname is not, just add sb_roadname and an '&'. 
        elif nb_roadname == "" and len(sb_roadname) > 0:
            header_str = header_str + sb_roadname + " & "
        # If sb_roadname is empty while nb_roadname is not just add nb_roadname and an '&'. 
        elif sb_roadname == "" and len(nb_roadname) > 0:
            header_str = header_str + nb_roadname
        # If both nb_roadname and sb_roadname are not empty and are different, then add both road names separated by a '/' and an '&' at the end. 
        else: 
            header_str = header_str + nb_roadname + " / " + sb_roadname + " & "

        # If both road names the same and not empty, add the road name to the header string so far. 
        if eb_roadname == wb_roadname and len(eb_roadname) > 0: 
            header_str = header_str + eb_roadname
        # If nb_roadname is empty while sb_roadname is not, just add sb_roadname. 
        elif eb_roadname == "" and len(wb_roadname) > 0:
            header_str = header_str + wb_roadname
        # If sb_roadname is empty while nb_roadname is not just add nb_roadname. 
        elif wb_roadname == "" and len(eb_roadname) > 0:
            header_str = header_str + eb_roadname
        # If both nb_roadname and sb_roadname are not empty and are different, then add both road names separated by a '/'. 
        else: 
            header_str = header_str + eb_roadname + " / " + wb_roadname

        # Populate intersection number box and add intersection number to the header string. 
        int_num = df.loc[df_row, 'Int. ID 1']
        header_str = "Intersection " + str(int_num) + ": " + header_str

        # Insert header string into the header and format. 
        ws[origin].value = header_str
        ws[origin].alignment = Alignment(horizontal='center', vertical='center')

    return    

def isValidHexaCode(input_str):
    '''
    PURPOSE/DESCRIPTION: Checks if a string is a valid hexadecimal color code. Borrowed and modified from https://www.geeksforgeeks.org/check-if-a-given-string-is-a-valid-hexadecimal-color-code-or-not/. 

    INPUT:
    input_str (String): The string to check if it is a valid hexadecimal color code. 

    OUTPUT:
    Either True or False.  
    '''
    str = input_str.lower()

    if (not(len(str) == 3 or len(str) == 6)):
        return False
 
    for i in range(0, len(str)):
        if (not((str[i] >= '0' and str[i] <= '9') or (str[i] >= 'a' and str[i] <= 'f') or (str[i] >= 'A' or str[i] <= 'F'))):
            return False
 
    return True

    
'''
****************** EXECUTING THE MAIN SCRIPT BELOW ******************
'''
# Get the current script directory. 
script_dir = os.path.dirname(os.path.abspath(__file__))

# Reference _data merge.csv into hte dataframe. 
csv_path = script_dir + "\\_data merge.csv"
df = pd.read_csv(csv_path)

# Front-fill the scenarios in the dataframe (or flash fill). 
df['Scenario'] = df['Scenario'].ffill()

# Make the selections for customizations here, GLOBAL VARIABLES FOR OPTIONS!
origin = 'B2'
main_display_height = 26    # Would recommend keeping this value. Recommend at least even number (code may not work well for odd) and at least 26. 
main_display_width = 24     # Would recommend keeping this value. Recommend at least even number (code may not work well for odd) and at least 24. 

# Obtain user input for header selection (yes or no). 
user_input_header = input("\nWould you like a header for your figures? Please enter Y or N: ")
while user_input_header.upper() != 'Y' and user_input_header.upper() != 'N': 
    user_input_header = input("\nSilly Goose, that wasn't a valid response. Would you like a header for your figures? Please enter Y or N: ")
if user_input_header.upper() == 'Y':
    header = True            
elif user_input_header.upper() == 'N':
    header = False 
else:
    raise Exception("Something went wrong if this error shows up.")

# If the user selects yes to header, set the header_height. 
if header:
    header_height = 2

# Obtain user input for main border selection (yes or no). 
user_input_main_border = input("\nWould you like a border for the main display area of your figures? Please enter Y or N: ")
while user_input_main_border.upper() != 'Y' and user_input_main_border.upper() != 'N': 
    user_input_main_border = input("\nSilly Goose, that wasn't a valid response. Would you like a border for the main display area of your figures? Please enter Y or N: ")
if user_input_main_border.upper() == 'Y':
    main_border = True            
elif user_input_main_border.upper() == 'N':
    main_border = False 
else:
    raise Exception("Something went wrong if this error shows up.")

# Obtain user input for cardinal directions selection (yes or no). 
user_input_cardinal_dirs = input("\nWould you like cardinal directions on your figures? Please enter Y or N: ")
while user_input_cardinal_dirs.upper() != 'Y' and user_input_cardinal_dirs.upper() != 'N': 
    user_input_cardinal_dirs = input("\nSilly Goose, that wasn't a valid response. Would you like cardinal directions on your figures? Please enter Y or N: ")
if user_input_cardinal_dirs.upper() == 'Y':
    cardinal_dirs = True            
elif user_input_cardinal_dirs.upper() == 'N':
    cardinal_dirs = False 
else:
    raise Exception("Something went wrong if this error shows up.")

# Obtain user input for intersection number box selection (yes or no). 
user_input_int_num_box = input("\nWould you like an intersection number box for your figures? Please enter Y or N: ")
while user_input_int_num_box.upper() != 'Y' and user_input_int_num_box.upper() != 'N': 
    user_input_int_num_box = input("\nSilly Goose, that wasn't a valid response. Would you like an intersection number box for your figures? Please enter Y or N: ")
if user_input_int_num_box.upper() == 'Y':
    int_num_box = True            
elif user_input_int_num_box.upper() == 'N':
    int_num_box = False 
else:
    raise Exception("Something went wrong if this error shows up.")

# Obtain user input for main display area background color; if empty string is provided, will use default of 'A6C9EC'. 
main_bkgd_color = 'A6C9EC'
user_input_main_bkgd_color = input("\nPlease input the hexadecimal color code of the main background color you'd like to use without the # (if you wish to use the default color, just hit Enter): ")
while not isValidHexaCode(user_input_main_bkgd_color) and user_input_main_bkgd_color != '': 
    user_input_main_bkgd_color = input("\nSilly Goose, that wasn't a valid response. Please input the hexadecimal color code of the main background color you'd like to use without the # (if you wish to use the default color, just hit Enter): ")
if user_input_main_bkgd_color != '':
    main_bkgd_color = user_input_main_bkgd_color

# Obtain user input for main border color, since user indicated they wanted a border; if empty string is provided, will use default of 'DAE9F8'. 
if main_border: 
    main_border_color = 'DAE9F8'      
    user_input_main_border_color = input("\nPlease input the hexadecimal color code of the main border color you'd like to use without the # (if you wish to use the default color, just hit Enter): ")
    while not isValidHexaCode(user_input_main_border_color) and user_input_main_border_color != '': 
        user_input_main_border_color = input("\nSilly Goose, that wasn't a valid response. Please input the hexadecimal color code of the main border color you'd like to use without the # (if you wish to use the default color, just hit Enter): ")
    if user_input_main_border_color != '':
        main_border_color = user_input_main_border_color

# Obtain user input for header color, since user indicated they wanted a header; if empty string is provided, will use default of '83CCEB'. 
if header: 
    header_color = '83CCEB'  
    user_input_header_color = input("\nPlease input the hexadecimal color code of the header color you'd like to use without the # (if you wish to use the default color, just hit Enter): ")
    while not isValidHexaCode(user_input_header_color) and user_input_header_color != '': 
        user_input_header_color = input("\nSilly Goose, that wasn't a valid response. Please input the hexadecimal color code of the header color you'd like to use without the # (if you wish to use the default color, just hit Enter): ")
    if user_input_header_color != '':
        header_color = user_input_header_color

# Obtain user input for intersection number box color, since user indicated they wanted a they wanted one; if empty string is provided, will use default of 'C6C9EC'. 
if int_num_box: 
    int_num_box_color = 'C6C9EC'  
    user_input_int_num_box_color = input("\nPlease input the hexadecimal color code of the intersection number box color you'd like to use without the # (if you wish to use the default color, just hit Enter): ")
    while not isValidHexaCode(user_input_int_num_box_color) and user_input_int_num_box_color != '': 
        user_input_int_num_box_color = input("\nSilly Goose, that wasn't a valid response. Please input the hexadecimal color code of the intersection number box color you'd like to use without the # (if you wish to use the default color, just hit Enter): ")
    if user_input_int_num_box_color != '':
        int_num_box_color = user_input_int_num_box_color

# Define gap (and "jump") between the figures in the worksheet. 
gap = 3
jump = main_display_height + gap 
if header:
    jump = jump + 2

# The relative path of where the images are stored. 
img_dir_path = '.\\PNG'

# Create the Excel workbook. 
wb = Workbook()

# Determine all of the unique conditions scenarios, create a worksheet for each one, and set the zoom to 115%; update progress printing to terminal.
progress = 0
progress_i = 1 / len(df.index)
unique_scenarios = df['Scenario'].unique()
for scenario in unique_scenarios: 
    ws = wb.create_sheet(scenario)
    ws.sheet_view.zoomScale = 115

    # Set the column widths. 
    setColumnWidths(ws, 2.6)

    # Split the origin's cell address (where to begin figure generation at) for ease of use and set curr_row to origin_row.  
    origin_col, origin_row = splitCellCoord(origin)
    curr_row = int(origin_row)

    # For each row in the dataframe (a row represents one intersection under one scenario)...
    for i in range(len(df.index)):
        
        # If the row is the scenario that we're currently iterating through...
        if df.loc[i, 'Scenario'] == scenario:

            # Set the local figure origin to the origin column and current row, then generate and populate the figure, and move curr_row to where the next figure will be created.
            local_fig_origin = origin_col + str(curr_row)
            generateFigure(ws, df, i, local_fig_origin)
            populateFigure(ws, df, i, local_fig_origin)
            curr_row = curr_row + jump

            progress = progress + progress_i
            print(str(math.ceil(progress * 100)) + '% complete!')

    


# Remove the default 'Sheet' worksheet and save the workbook. 
wb.remove(wb['Sheet'])
wb.save(fr'{script_dir}\Figures.xlsx')

